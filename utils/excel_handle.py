# -*- coding:utf-8 -*-
import io
import os.path

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles.colors import BLACK
from openpyxl.utils import get_column_letter


class ExcelHandle:
    """解析excel文件"""

    def __init__(self, filename, sheet_name=''):
        """
        excel读取和写入
        :param filename:
        :param sheet_name:
        """
        try:
            self.sheet_name = sheet_name
            self.filename = filename
            self.__wb = load_workbook(self.filename)  # 打开excel
        except FileNotFoundError as e:
            self.__wb = Workbook()

    def get_ws_by_sheet(self):
        if self.sheet_name:
            ws = self.__wb[self.sheet_name]
        else:
            ws = self.__wb[self.__wb.sheetnames[0]]
        return ws

    def get_max_row_num(self):
        """获取最大行号"""
        ws = self.get_ws_by_sheet()
        max_row_num = ws.max_row
        return max_row_num

    def get_max_column_num(self):
        """获取最大列号"""
        ws = self.get_ws_by_sheet()
        max_column = ws.max_column
        return max_column

    def get_head(self):
        """
        获取表头
        :return:
        """
        ws = self.get_ws_by_sheet()
        head = next(ws.iter_rows(max_row=1, values_only=True))
        return head

    def get_cell_value(self, coordinate=None, row=None, column=None):
        ws = self.get_ws_by_sheet()
        """获取指定单元格的数据"""
        if coordinate is not None:
            try:
                return ws[coordinate].value
            except Exception as e:
                raise e
        elif coordinate is None and row is not None and column is not None:
            if isinstance(row, int) and isinstance(column, int):
                return ws.cell(row=row, column=column).value
            else:
                raise TypeError('row and column must be type int')
        else:
            raise Exception("Insufficient Coordinate of cell!")

    def get_row_value(self, row):
        """获取某一行的数据"""
        ws = self.get_ws_by_sheet()
        col_num = self.get_max_column_num()
        row_value = []
        if isinstance(row, int):
            for col in range(1, col_num + 1):
                values_row = ws.cell(row, col).value
                row_value.append(values_row)
            return row_value
        else:
            raise TypeError('row must be type int')

    def get_column_value(self, column):
        """获取某一列数据"""
        ws = self.get_ws_by_sheet()
        row_num = self.get_max_row_num()
        column_value = []
        if isinstance(column, int):
            for row in range(1, row_num + 1):
                values_column = ws.cell(row, column).value
                column_value.append(values_column)
            return column_value
        else:
            raise TypeError('column must be type int')

    def get_all_value(self, exclude_head=True):
        """
        获取指定表单的所有数据
        :param exclude_head:是否去掉表头
        :return:
        """
        ws = self.get_ws_by_sheet()
        min_row = 2 if exclude_head else 1
        row = ws.iter_rows(min_row=min_row, max_row=ws.max_row, values_only=True)
        values = []
        for row_tuple in row:
            if self.is_row_none(row_tuple):
                continue
            value_list = []
            for value in row_tuple:
                value_list.append(value)
            values.append(value_list)
        return values

    def is_row_none(self, row):
        """
        判断当前行是否全为None
        :param row:
        :return:
        """
        return not list(filter(None, row))

    def get_excel_title(self):
        """获取sheet表头"""
        ws = self.get_ws_by_sheet()
        title_key = tuple(ws.iter_rows(max_row=1, values_only=True))[0]
        return title_key

    def get_list_dict_all_value(self, exclude_head=True):
        """
        返回字典列表
        :param exclude_head:是否去掉表头
        :return:
        """
        head = self.get_head()
        all_values = self.get_all_value(exclude_head)
        value_list = []
        for value in all_values:
            value_list.append(dict(zip(head, value)))
        return value_list

    def write_cell_excel(self, row, column, value='', bold=True, color=BLACK):
        if isinstance(row, int) and isinstance(column, int):
            ws = self.get_ws_by_sheet()
            cell = ws.cell(row, column)
            cell.font = Font(color=color, bold=bold)
            cell.value = value
        else:
            raise TypeError('row and column must be type int')

    def write_cell(self, row, column, value=None, bold=True, color=BLACK):
        ws = self.get_ws_by_sheet()
        if isinstance(row, int) and isinstance(column, int):
            try:
                cell_obj = ws.cell(row, column)
                cell_obj.font = Font(color=color, bold=bold)
                cell_obj.value = value
                self.__wb.save(self.filename)
            except Exception as e:
                raise e
        else:
            raise TypeError('row and column must be type int')

    def save_excel_io(self):
        """
        文件写入流
        :return:
        """
        xlsx_file = io.BytesIO()
        self.__wb.save(xlsx_file)
        return xlsx_file

    def auto_alignment(self, row, col):
        """
        自动换行
        :param row:
        :param col:
        :return:
        """
        ws = self.get_ws_by_sheet()
        ws.cell(row, col).alignment = Alignment(wrapText=True)

    def adjust_col_alignment(self, col):
        """
        自适应列宽
        :param col:
        :return:
        """
        max_length = 0
        ws = self.get_ws_by_sheet()
        for cell in ws[get_column_letter(col)]:
            length = 0
            for char in str(cell.value):
                if ord(char) > 255:
                    # 中文长度自增2
                    length += 2
                else:
                    length += 1.2
            if length > max_length:
                max_length = length
        # 调整列宽
        ws.column_dimensions[get_column_letter(col)].width = max_length


if __name__ == '__main__':
    from configs.dir_path import data_path
    excel = ExcelHandle(filename=os.path.join(data_path, 'testcase.xlsx'))
    print(excel.get_list_dict_all_value())